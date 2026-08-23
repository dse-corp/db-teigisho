"""Human-readable HTML, Excel, and PDF renderers."""

from __future__ import annotations

import hashlib
import html
import json
import re
from base64 import b64encode
from collections.abc import Mapping
from datetime import UTC, datetime
from importlib.resources import as_file, files
from io import BytesIO
from pathlib import Path
from typing import Any

from jinja2 import Environment, PackageLoader, select_autoescape
from openpyxl import Workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.properties import PageSetupProperties
from openpyxl.worksheet.worksheet import Worksheet
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import (
    Image as ReportlabImage,
)
from reportlab.platypus import (
    PageBreak,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
    XPreformatted,
)

from db_teigisho import __version__
from db_teigisho.diagram_render import (
    DEFAULT_ER_DIAGRAM_MODES,
    RenderedErDiagram,
    render_er_diagrams,
)
from db_teigisho.er import ColumnDisplayMode, render_mermaid
from db_teigisho.loader import load_definition
from db_teigisho.models import DatabaseDefinition, SqlObjectDefinition, TableDefinition

_BRAND = "124E78"
_LIGHT = "E8F2F8"
_PDF_FONT = "NotoSansJP"
_THIN_SIDE = Side(style="thin", color="D8DEE5")
_CELL_BORDER = Border(
    left=_THIN_SIDE,
    right=_THIN_SIDE,
    top=_THIN_SIDE,
    bottom=_THIN_SIDE,
)
_EXCEL_FORMULA_PREFIXES = ("=", "+", "-", "@")


def _ensure_parent(path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)


def _excel_value(value: object) -> Any:
    """Force formula-like user input to remain text in generated workbooks."""

    if isinstance(value, str) and value.startswith(_EXCEL_FORMULA_PREFIXES):
        return f"'{value}"
    return value


def _svg_data_uri(svg: bytes) -> str:
    return f"data:image/svg+xml;base64,{b64encode(svg).decode('ascii')}"


def _all_er_diagrams(
    definition: DatabaseDefinition,
    diagrams: Mapping[ColumnDisplayMode, RenderedErDiagram] | None,
) -> Mapping[ColumnDisplayMode, RenderedErDiagram]:
    return diagrams if diagrams is not None else render_er_diagrams(definition)


def _default_er_diagram(
    definition: DatabaseDefinition,
    diagrams: Mapping[ColumnDisplayMode, RenderedErDiagram] | None,
) -> RenderedErDiagram:
    if diagrams is not None:
        return diagrams["all"]
    return render_er_diagrams(definition, modes=("all",))["all"]


def render_html(
    definition: DatabaseDefinition,
    output: Path,
    diagrams: Mapping[ColumnDisplayMode, RenderedErDiagram] | None = None,
) -> None:
    """Render a standalone, responsive HTML definition."""

    _ensure_parent(output)
    environment = Environment(
        loader=PackageLoader("db_teigisho", "templates"),
        autoescape=select_autoescape(("html", "xml")),
        trim_blocks=True,
        lstrip_blocks=True,
    )
    rendered_diagrams = _all_er_diagrams(definition, diagrams)
    template = environment.get_template("definition.html.j2")
    output.write_text(
        template.render(
            definition=definition,
            er_diagram_images={
                mode: _svg_data_uri(rendered_diagrams[mode].svg)
                for mode in DEFAULT_ER_DIAGRAM_MODES
            },
            tool_version=__version__,
        ),
        encoding="utf-8",
    )


def _style_title(sheet: Worksheet, title: str, end_column: int = 2) -> None:
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=end_column)
    cell = sheet.cell(1, 1, _excel_value(title))
    cell.font = Font(color="FFFFFF", bold=True, size=16)
    cell.fill = PatternFill("solid", fgColor=_BRAND)
    cell.alignment = Alignment(vertical="center")
    sheet.row_dimensions[1].height = 28


def _style_header(sheet: Worksheet, row: int, columns: int) -> None:
    for cell in sheet[row][:columns]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = PatternFill("solid", fgColor=_BRAND)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _CELL_BORDER


def _fit_columns(sheet: Worksheet, minimum: int = 9, maximum: int = 48) -> None:
    for column_cells in sheet.columns:
        column_number = column_cells[0].column
        if not isinstance(column_number, int):
            continue
        letter = get_column_letter(column_number)
        size = max(
            (len(str(cell.value)) for cell in column_cells if cell.value is not None),
            default=minimum,
        )
        sheet.column_dimensions[letter].width = min(max(size + 2, minimum), maximum)


def _document_sheet(workbook: Workbook, definition: DatabaseDefinition) -> None:
    sheet = workbook.active
    assert isinstance(sheet, Worksheet)
    sheet.title = "文書情報"
    sheet.sheet_view.showGridLines = False
    _style_title(sheet, f"{definition.document.system_name} データベース定義書")
    values = [
        ("プロジェクト番号", definition.document.project_number),
        ("システム名", definition.document.system_name),
        ("サブシステム名", definition.document.subsystem_name),
        ("作成日時", definition.document.created_at.isoformat()),
        ("変更日時", definition.document.updated_at.isoformat()),
        ("DBMS名", definition.database.dbms_name),
        ("DBMSバージョン", definition.database.dbms_version),
        ("サーバー名", definition.database.server_name),
        ("ポート", definition.database.port),
        ("データベース名", definition.database.database_name),
        ("スキーマ名", definition.database.schema_name),
        ("Collation", definition.database.collation),
        ("フォーマットバージョン", definition.format_version),
    ]
    for row, (label, value) in enumerate(values, start=2):
        sheet.cell(row, 1, label).font = Font(bold=True, color="4A5968")
        sheet.cell(row, 1).fill = PatternFill("solid", fgColor="F1F4F7")
        sheet.cell(row, 2, _excel_value(value if value is not None else ""))
    sheet.column_dimensions["A"].width = 24
    sheet.column_dimensions["B"].width = 58
    sheet.freeze_panes = "A2"
    sheet.page_setup.orientation = "portrait"
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 1
    sheet.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    sheet.print_area = "A1:B14"


def _safe_sheet_title(raw: str, existing: set[str]) -> str:
    base = re.sub(r"[\\/*?:\[\]]", "_", raw).strip("'") or "table"
    base = base[:31]
    candidate = base
    counter = 2
    while candidate.casefold() in existing:
        suffix = f"_{counter}"
        candidate = f"{base[: 31 - len(suffix)]}{suffix}"
        counter += 1
    existing.add(candidate.casefold())
    return candidate


def _table_list_sheet(workbook: Workbook, definition: DatabaseDefinition) -> None:
    sheet = workbook.create_sheet("テーブル一覧")
    sheet.sheet_view.showGridLines = False
    headers = [
        "No.",
        "テーブル物理名",
        "テーブル論理名",
        "説明",
        "カラム数",
        "インデックス数",
        "外部キー数",
    ]
    _style_title(sheet, f"テーブル一覧 ({len(definition.tables)})", len(headers))
    for column, header in enumerate(headers, start=1):
        sheet.cell(3, column, header)
    _style_header(sheet, 3, len(headers))
    for row, table in enumerate(definition.tables, start=4):
        values: list[object] = [
            row - 3,
            table.physical_name,
            table.logical_name,
            table.description or "",
            len(table.columns),
            len(table.indexes),
            len(table.foreign_keys),
        ]
        for number, value in enumerate(values, start=1):
            cell = sheet.cell(row, number, _excel_value(value))
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = _CELL_BORDER
        if row % 2 == 1:
            for cell in sheet[row][: len(headers)]:
                cell.fill = PatternFill("solid", fgColor="F6F8FA")

    widths = [8, 28, 28, 54, 14, 16, 14]
    for column, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(column)].width = width
    sheet.freeze_panes = "A4"
    sheet.auto_filter.ref = f"A3:G{max(3, 3 + len(definition.tables))}"
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    sheet.print_title_rows = "1:3"
    sheet.print_area = sheet.dimensions


def _er_sheet(workbook: Workbook, diagram: RenderedErDiagram) -> None:
    sheet = workbook.create_sheet("ER図")
    sheet.sheet_view.showGridLines = False
    _style_title(sheet, "ER図（全カラム）", 16)
    sheet.cell(2, 1, "PK/FKとカーディナリティを含むER図").font = Font(bold=True, color="4A5968")
    image = OpenpyxlImage(BytesIO(diagram.png))
    maximum_width = 1800
    if image.width > maximum_width:
        scale = maximum_width / image.width
        image.width = maximum_width
        image.height = int(image.height * scale)
    sheet.add_image(image, "A3")
    for column in range(1, 17):
        sheet.column_dimensions[get_column_letter(column)].width = 12
    for row in range(3, 46):
        sheet.row_dimensions[row].height = 18
    sheet.freeze_panes = "A3"
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    sheet.print_area = "A1:P45"


def _table_sheet(workbook: Workbook, table: TableDefinition, title: str) -> None:
    sheet = workbook.create_sheet(title)
    sheet.sheet_view.showGridLines = False
    headers = [
        "No.",
        "カラム物理名",
        "カラム論理名",
        "型",
        "長さ",
        "Scale",
        "デフォルト値",
        "NN",
        "Unique",
        "PK",
        "説明",
    ]
    _style_title(sheet, f"{table.logical_name} ({table.physical_name})", len(headers))
    sheet.cell(2, 1, "説明").font = Font(bold=True)
    sheet.merge_cells(start_row=2, start_column=2, end_row=2, end_column=len(headers))
    sheet.cell(2, 2, _excel_value(table.description or ""))
    for column, header in enumerate(headers, start=1):
        sheet.cell(4, column, header)
    _style_header(sheet, 4, len(headers))
    for row, column_definition in enumerate(table.columns, start=5):
        values: list[Any] = [
            row - 4,
            column_definition.physical_name,
            column_definition.logical_name,
            column_definition.data_type,
            column_definition.length,
            column_definition.scale,
            column_definition.default,
            "✓" if column_definition.not_null else "",
            "✓" if column_definition.unique else "",
            "✓" if column_definition.primary_key else "",
            column_definition.description,
        ]
        for number, value in enumerate(values, start=1):
            cell = sheet.cell(
                row,
                number,
                _excel_value(value if value is not None else ""),
            )
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = _CELL_BORDER
        if row % 2 == 0:
            for cell in sheet[row][: len(headers)]:
                cell.fill = PatternFill("solid", fgColor="F6F8FA")

    next_row = 5 + len(table.columns) + 2
    sheet.cell(next_row, 1, "インデックス").font = Font(bold=True, color=_BRAND, size=13)
    index_headers = ["名前", "種別", "Unique", "キー列", "付加列", "条件"]
    for column, header in enumerate(index_headers, start=1):
        sheet.cell(next_row + 1, column, header)
    _style_header(sheet, next_row + 1, len(index_headers))
    for row, index in enumerate(table.indexes, start=next_row + 2):
        values = [
            index.name,
            index.type or "",
            "✓" if index.unique else "",
            ", ".join(f"{item.name} {item.order}" for item in index.columns),
            ", ".join(index.include_columns),
            index.where or "",
        ]
        for number, value in enumerate(values, start=1):
            sheet.cell(row, number, _excel_value(value)).border = _CELL_BORDER

    fk_row = next_row + max(len(table.indexes), 1) + 4
    sheet.cell(fk_row, 1, "外部キー").font = Font(bold=True, color=_BRAND, size=13)
    fk_headers = ["名前", "参照元", "参照先", "ON UPDATE", "ON DELETE", "Deferrable"]
    for column, header in enumerate(fk_headers, start=1):
        sheet.cell(fk_row + 1, column, header)
    _style_header(sheet, fk_row + 1, len(fk_headers))
    for row, foreign_key in enumerate(table.foreign_keys, start=fk_row + 2):
        values = [
            foreign_key.name,
            ", ".join(foreign_key.columns),
            f"{foreign_key.referenced_table}({', '.join(foreign_key.referenced_columns)})",
            foreign_key.on_update,
            foreign_key.on_delete,
            "✓" if foreign_key.deferrable else "",
        ]
        for number, value in enumerate(values, start=1):
            sheet.cell(row, number, _excel_value(value)).border = _CELL_BORDER
    _fit_columns(sheet)
    sheet.freeze_panes = "A5"
    sheet.auto_filter.ref = f"A4:{get_column_letter(len(headers))}{4 + len(table.columns)}"
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    sheet.print_title_rows = "1:4"
    sheet.print_area = sheet.dimensions


def _sql_sheet(workbook: Workbook, title: str, objects: list[SqlObjectDefinition]) -> None:
    sheet = workbook.create_sheet(title)
    sheet.sheet_view.showGridLines = False
    headers = ["物理名", "論理名", "説明", "SQL"]
    _style_title(sheet, title, len(headers))
    for column, header in enumerate(headers, start=1):
        sheet.cell(3, column, header)
    _style_header(sheet, 3, len(headers))
    for row, item in enumerate(objects, start=4):
        values = [item.physical_name, item.logical_name, item.description or "", item.sql]
        for number, value in enumerate(values, start=1):
            cell = sheet.cell(row, number, _excel_value(value))
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = _CELL_BORDER
        sheet.row_dimensions[row].height = min(120, 18 * max(1, item.sql.count("\n") + 1))
    _fit_columns(sheet)
    sheet.column_dimensions["D"].width = 90
    sheet.freeze_panes = "A4"
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    sheet.print_title_rows = "1:3"
    sheet.print_area = sheet.dimensions


def render_xlsx(
    definition: DatabaseDefinition,
    output: Path,
    diagrams: Mapping[ColumnDisplayMode, RenderedErDiagram] | None = None,
) -> None:
    """Render an Excel workbook optimized for review and printing."""

    _ensure_parent(output)
    workbook = Workbook()
    _document_sheet(workbook, definition)
    _table_list_sheet(workbook, definition)
    _er_sheet(workbook, _default_er_diagram(definition, diagrams))
    existing = {"文書情報".casefold(), "テーブル一覧".casefold(), "ER図".casefold()}
    for table in definition.tables:
        _table_sheet(workbook, table, _safe_sheet_title(table.physical_name, existing))
    _sql_sheet(workbook, "ビュー", definition.views)
    _sql_sheet(workbook, "ストアドプロシージャ", definition.stored_procedures)
    workbook.save(output)


def _pdf_styles() -> dict[str, ParagraphStyle]:
    font_resource = files("db_teigisho").joinpath("assets/NotoSansJP[wght].ttf")
    with as_file(font_resource) as font_path:
        pdfmetrics.registerFont(TTFont(_PDF_FONT, str(font_path)))
    samples = getSampleStyleSheet()
    return {
        "title": ParagraphStyle(
            "JapaneseTitle",
            parent=samples["Title"],
            fontName=_PDF_FONT,
            fontSize=21,
            leading=28,
            textColor=colors.HexColor(f"#{_BRAND}"),
            alignment=TA_CENTER,
        ),
        "h1": ParagraphStyle(
            "JapaneseHeading1",
            parent=samples["Heading1"],
            fontName=_PDF_FONT,
            fontSize=15,
            leading=20,
            textColor=colors.HexColor(f"#{_BRAND}"),
            spaceBefore=8,
            spaceAfter=8,
        ),
        "h2": ParagraphStyle(
            "JapaneseHeading2",
            parent=samples["Heading2"],
            fontName=_PDF_FONT,
            fontSize=12,
            leading=16,
            textColor=colors.HexColor(f"#{_BRAND}"),
            spaceBefore=8,
            spaceAfter=5,
        ),
        "body": ParagraphStyle(
            "JapaneseBody",
            parent=samples["BodyText"],
            fontName=_PDF_FONT,
            fontSize=8,
            leading=11,
        ),
        "header": ParagraphStyle(
            "JapaneseTableHeader",
            parent=samples["BodyText"],
            fontName=_PDF_FONT,
            fontSize=7,
            leading=9,
            textColor=colors.white,
        ),
        "small": ParagraphStyle(
            "JapaneseSmall",
            parent=samples["BodyText"],
            fontName=_PDF_FONT,
            fontSize=6.5,
            leading=8,
        ),
        "sql": ParagraphStyle(
            "JapaneseSql",
            parent=samples["Code"],
            fontName=_PDF_FONT,
            fontSize=7,
            leading=9,
            leftIndent=4,
            rightIndent=4,
            borderColor=colors.HexColor("#C8D0D8"),
            borderWidth=0.5,
            borderPadding=6,
            backColor=colors.HexColor("#F4F6F8"),
        ),
    }


def _paragraph(value: object, style: ParagraphStyle) -> Paragraph:
    text = "" if value is None else str(value)
    return Paragraph(html.escape(text).replace("\n", "<br/>"), style)


def _pdf_table(
    rows: list[list[object]],
    styles: dict[str, ParagraphStyle],
    widths: list[float] | None = None,
) -> Table:
    rendered = [
        [
            _paragraph(value, styles["small"] if row_number else styles["header"])
            for value in row
        ]
        for row_number, row in enumerate(rows)
    ]
    table = Table(rendered, colWidths=widths, repeatRows=1, hAlign="LEFT")
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(f"#{_BRAND}")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, -1), _PDF_FONT),
                ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#C8D0D8")),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F6F8FA")]),
                ("LEFTPADDING", (0, 0), (-1, -1), 3),
                ("RIGHTPADDING", (0, 0), (-1, -1), 3),
                ("TOPPADDING", (0, 0), (-1, -1), 3),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ]
        )
    )
    return table


def _page_number(canvas: Any, document: Any) -> None:
    canvas.saveState()
    canvas.setFont(_PDF_FONT, 7)
    canvas.setFillColor(colors.HexColor("#5D6D7E"))
    canvas.drawRightString(landscape(A4)[0] - 14 * mm, 8 * mm, f"{document.page}")
    canvas.restoreState()


def _pdf_er_diagram(diagram: RenderedErDiagram) -> ReportlabImage:
    image = ReportlabImage(BytesIO(diagram.png))
    maximum_width = landscape(A4)[0] - 28 * mm
    maximum_height = landscape(A4)[1] - 48 * mm
    scale = min(maximum_width / image.imageWidth, maximum_height / image.imageHeight, 1)
    image.drawWidth = image.imageWidth * scale
    image.drawHeight = image.imageHeight * scale
    return image


def render_pdf(
    definition: DatabaseDefinition,
    output: Path,
    diagrams: Mapping[ColumnDisplayMode, RenderedErDiagram] | None = None,
) -> None:
    """Render a Japanese-capable review PDF."""

    _ensure_parent(output)
    styles = _pdf_styles()
    document = SimpleDocTemplate(
        str(output),
        pagesize=landscape(A4),
        leftMargin=14 * mm,
        rightMargin=14 * mm,
        topMargin=13 * mm,
        bottomMargin=13 * mm,
        title=f"{definition.document.system_name} データベース定義書",
        author="db-teigisho",
    )
    story: list[Any] = [
        Spacer(1, 18 * mm),
        Paragraph(
            f"{html.escape(definition.document.system_name)}<br/>データベース定義書",
            styles["title"],
        ),
        Spacer(1, 8 * mm),
        _pdf_table(
            [
                ["項目", "値"],
                ["プロジェクト番号", definition.document.project_number],
                ["サブシステム名", definition.document.subsystem_name or "—"],
                ["作成日時", definition.document.created_at.isoformat()],
                ["変更日時", definition.document.updated_at.isoformat()],
                ["DBMS", f"{definition.database.dbms_name} {definition.database.dbms_version}"],
                [
                    "接続先",
                    f"{definition.database.server_name or '—'}:{definition.database.port or '—'}",
                ],
                [
                    "Database / Schema",
                    f"{definition.database.database_name} / {definition.database.schema_name}",
                ],
                ["Collation", definition.database.collation or "—"],
            ],
            styles,
            [48 * mm, 120 * mm],
        ),
    ]

    story.extend(
        [
            PageBreak(),
            Paragraph("ER図（全カラム）", styles["h1"]),
            _pdf_er_diagram(_default_er_diagram(definition, diagrams)),
            PageBreak(),
            Paragraph(f"テーブル一覧（{len(definition.tables)}）", styles["h1"]),
        ]
    )
    if definition.tables:
        table_list_rows: list[list[object]] = [
            [
                "No.",
                "テーブル物理名",
                "テーブル論理名",
                "説明",
                "カラム数",
                "インデックス数",
                "外部キー数",
            ]
        ]
        for number, table in enumerate(definition.tables, start=1):
            table_list_rows.append(
                [
                    number,
                    table.physical_name,
                    table.logical_name,
                    table.description or "",
                    len(table.columns),
                    len(table.indexes),
                    len(table.foreign_keys),
                ]
            )
        story.append(
            _pdf_table(
                table_list_rows,
                styles,
                [10 * mm, 38 * mm, 35 * mm, 94 * mm, 18 * mm, 22 * mm, 20 * mm],
            )
        )
    else:
        story.append(Paragraph("テーブル定義なし", styles["body"]))

    for table in definition.tables:
        story.extend(
            [
                PageBreak(),
                Paragraph(
                    f"{html.escape(table.logical_name)} ({html.escape(table.physical_name)})",
                    styles["h1"],
                ),
            ]
        )
        if table.description:
            story.extend([_paragraph(table.description, styles["body"]), Spacer(1, 3 * mm)])
        column_rows: list[list[object]] = [
            ["No.", "物理名", "論理名", "型", "長さ", "Scale", "Default", "NN", "UQ", "PK", "説明"]
        ]
        for number, column in enumerate(table.columns, start=1):
            column_rows.append(
                [
                    number,
                    column.physical_name,
                    column.logical_name,
                    column.data_type,
                    column.length or "",
                    "" if column.scale is None else column.scale,
                    "" if column.default is None else column.default,
                    "✓" if column.not_null else "",
                    "✓" if column.unique else "",
                    "✓" if column.primary_key else "",
                    column.description or "",
                ]
            )
        story.extend(
            [
                _pdf_table(
                    column_rows,
                    styles,
                    [
                        9 * mm,
                        30 * mm,
                        31 * mm,
                        22 * mm,
                        13 * mm,
                        12 * mm,
                        23 * mm,
                        10 * mm,
                        10 * mm,
                        10 * mm,
                        45 * mm,
                    ],
                ),
                Spacer(1, 5 * mm),
                Paragraph("インデックス", styles["h2"]),
            ]
        )
        index_rows: list[list[object]] = [["名前", "種別", "Unique", "キー列", "付加列", "条件"]]
        index_rows.extend(
            [
                index.name,
                index.type or "",
                "✓" if index.unique else "",
                ", ".join(f"{item.name} {item.order}" for item in index.columns),
                ", ".join(index.include_columns),
                index.where or "",
            ]
            for index in table.indexes
        )
        if table.indexes:
            story.append(_pdf_table(index_rows, styles))
        else:
            story.append(Paragraph("定義なし", styles["body"]))
        story.extend([Spacer(1, 4 * mm), Paragraph("外部キー", styles["h2"])])
        fk_rows: list[list[object]] = [
            ["名前", "参照元", "参照先", "ON UPDATE", "ON DELETE", "Deferrable"]
        ]
        fk_rows.extend(
            [
                foreign_key.name,
                ", ".join(foreign_key.columns),
                f"{foreign_key.referenced_table}({', '.join(foreign_key.referenced_columns)})",
                foreign_key.on_update,
                foreign_key.on_delete,
                "✓" if foreign_key.deferrable else "",
            ]
            for foreign_key in table.foreign_keys
        )
        if table.foreign_keys:
            story.append(_pdf_table(fk_rows, styles))
        else:
            story.append(Paragraph("定義なし", styles["body"]))

    def add_sql_section(title: str, objects: list[SqlObjectDefinition]) -> None:
        story.extend([PageBreak(), Paragraph(title, styles["h1"])])
        if not objects:
            story.append(Paragraph("定義なし", styles["body"]))
        for item in objects:
            story.append(
                Paragraph(
                    f"{html.escape(item.logical_name)} ({html.escape(item.physical_name)})",
                    styles["h2"],
                )
            )
            if item.description:
                story.append(_paragraph(item.description, styles["body"]))
            story.extend(
                [
                    Spacer(1, 2 * mm),
                    XPreformatted(html.escape(item.sql), styles["sql"]),
                    Spacer(1, 4 * mm),
                ]
            )

    add_sql_section("ビュー", definition.views)
    add_sql_section("ストアドプロシージャ", definition.stored_procedures)
    document.build(story, onFirstPage=_page_number, onLaterPages=_page_number)


def _sha256(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def build_artifacts(source: Path, output_dir: Path) -> Path:
    """Validate once, render every format, and write a checksum manifest."""

    definition = load_definition(source)
    output_dir.mkdir(parents=True, exist_ok=True)
    stem = source.stem
    artifacts = {
        "html": output_dir / f"{stem}.html",
        "xlsx": output_dir / f"{stem}.xlsx",
        "pdf": output_dir / f"{stem}.pdf",
        "mermaid": output_dir / f"{stem}.mmd",
        "svg": output_dir / f"{stem}.svg",
        "png": output_dir / f"{stem}.png",
    }
    diagrams = render_er_diagrams(definition)
    render_html(definition, artifacts["html"], diagrams)
    render_xlsx(definition, artifacts["xlsx"], diagrams)
    render_pdf(definition, artifacts["pdf"], diagrams)
    artifacts["mermaid"].write_text(render_mermaid(definition), encoding="utf-8")
    artifacts["svg"].write_bytes(diagrams["all"].svg)
    artifacts["png"].write_bytes(diagrams["all"].png)
    manifest = {
        "format_version": "1.0",
        "generated_at": datetime.now(UTC).isoformat(),
        "generator": {"name": "db-teigisho", "version": __version__},
        "source": {"file": source.name, "sha256": _sha256(source)},
        "artifacts": [
            {
                "format": format_name,
                "file": path.name,
                "sha256": _sha256(path),
                "bytes": path.stat().st_size,
            }
            for format_name, path in artifacts.items()
        ],
    }
    manifest_path = output_dir / "manifest.json"
    manifest_path.write_text(
        json.dumps(manifest, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    return manifest_path
