from __future__ import annotations

import hashlib
import json
import re
from copy import deepcopy
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from db_teigisho.diagram_render import DEFAULT_ER_DIAGRAM_MODES, RenderedErDiagram
from db_teigisho.loader import load_definition
from db_teigisho.models import DatabaseDefinition
from db_teigisho.render import build_artifacts, render_html, render_pdf, render_xlsx


def test_renders_self_contained_html(definition_file: Path, tmp_path: Path) -> None:
    document = load_definition(definition_file)
    document.tables[0].description = "顧客の基本情報を管理する。"
    output = tmp_path / "definition.html"

    render_html(document, output)

    html = output.read_text(encoding="utf-8")
    assert "<!doctype html>" in html.lower()
    assert "受注管理システム" in html
    assert "orders" in html
    assert "customer_order_totals" in html
    assert "<style>" in html
    assert '<section id="table-list">' in html
    assert "テーブル一覧（2）" in html
    assert '<section id="er-diagram">' in html
    assert 'data-er-mode="all"' in html
    assert 'data-er-mode="keys"' in html
    assert 'data-er-mode="tables"' in html
    assert "data:image/svg+xml;base64," in html
    assert "全カラム" in html
    assert ".er-diagram-image[hidden] { display: none; }" in html
    assert "カラム数" in html
    assert "顧客の基本情報を管理する。" in html


def test_renders_interactive_er_viewer_controls_and_extension_contract(
    definition_file: Path, tmp_path: Path
) -> None:
    document = load_definition(definition_file)
    diagram = RenderedErDiagram(svg=b"<svg><text>static fallback</text></svg>", png=b"")
    diagrams = {mode: diagram for mode in DEFAULT_ER_DIAGRAM_MODES}
    output = tmp_path / "definition.html"

    render_html(document, output, diagrams)

    html = output.read_text(encoding="utf-8")
    assert 'id="dbdef-er-viewer"' in html
    assert 'data-er-action="zoom-in"' in html
    assert 'data-er-action="zoom-out"' in html
    assert 'data-er-action="fit"' in html
    assert 'id="dbdef-er-zoom-level"' in html
    assert 'class="er-diagram-fallback er-diagram-image"' in html
    assert "static fallback" not in html
    assert "window.dbdefErViewer" in html
    assert "setNodePosition" in html
    assert "setNodePositions" in html
    assert "redrawEdges" in html
    assert "setEdgePathRenderer" in html
    assert "dbdef:er-node-position-change" in html
    assert "dbdef:er-edges-redrawn" in html
    assert "@media print" in html
    assert re.search(r'(?:src|href)="https?://', html) is None


def test_embeds_restorable_graph_json_without_terminating_the_script_element(
    valid_definition: dict[str, Any], tmp_path: Path
) -> None:
    data = deepcopy(valid_definition)
    special_value = '</script> "quoted"\n日本語 & <tag>'
    data["tables"][0]["description"] = special_value
    data["tables"][0]["columns"][0]["logical_name"] = special_value
    definition = DatabaseDefinition.model_validate(data)
    diagram = RenderedErDiagram(svg=b"<svg></svg>", png=b"")
    diagrams = {mode: diagram for mode in DEFAULT_ER_DIAGRAM_MODES}
    output = tmp_path / "definition.html"

    render_html(definition, output, diagrams)

    html = output.read_text(encoding="utf-8")
    match = re.search(
        r'<script id="dbdef-er-graph" type="application/json">(.*?)</script>',
        html,
        re.DOTALL,
    )
    assert match is not None
    payload = match.group(1)
    assert "</script>" not in payload
    graph = json.loads(payload)
    assert graph["tables"][0]["description"] == special_value
    assert graph["tables"][0]["columns"][0]["logical_name"] == special_value


def test_renders_formatted_xlsx(definition_file: Path, tmp_path: Path) -> None:
    document = load_definition(definition_file)
    output = tmp_path / "definition.xlsx"

    render_xlsx(document, output)

    workbook = load_workbook(output, read_only=False, data_only=True)
    assert workbook.sheetnames == [
        "文書情報",
        "テーブル一覧",
        "ER図",
        "customers",
        "orders",
        "ビュー",
        "ストアドプロシージャ",
    ]
    assert workbook["文書情報"]["B2"].value == "PJ-001"
    assert workbook["テーブル一覧"]["B4"].value == "customers"
    assert workbook["テーブル一覧"]["C5"].value == "受注"
    assert workbook["テーブル一覧"]["E4"].value == 2
    assert workbook["テーブル一覧"]["F5"].value == 1
    assert workbook["テーブル一覧"]["G4"].value == 0
    assert workbook["テーブル一覧"].freeze_panes == "A4"
    assert len(workbook["ER図"]._images) == 1
    assert workbook["orders"].freeze_panes == "A5"
    assert workbook["orders"]["B5"].value == "order_id"
    assert workbook["orders"].sheet_view.showGridLines is False
    assert workbook["orders"].page_setup.orientation == "landscape"
    assert workbook["orders"].page_setup.fitToWidth == 1


def test_xlsx_keeps_formula_like_user_input_as_text(
    definition_file: Path, tmp_path: Path
) -> None:
    document = load_definition(definition_file)
    document.views[0].sql = '=HYPERLINK("https://example.test", "click")'
    output = tmp_path / "definition.xlsx"

    render_xlsx(document, output)

    workbook = load_workbook(output, read_only=False, data_only=False)
    cell = workbook["ビュー"]["D4"]
    assert cell.data_type == "s"
    assert cell.value == "'=HYPERLINK(\"https://example.test\", \"click\")"


def test_renders_a_valid_pdf(definition_file: Path, tmp_path: Path) -> None:
    document = load_definition(definition_file)
    document.views[0].sql = "SELECT * FROM orders WHERE total_amount < 100 AND note <> '<tag>';"
    output = tmp_path / "definition.pdf"

    render_pdf(document, output)

    content = output.read_bytes()
    assert content.startswith(b"%PDF-")
    assert len(content) > 3_000
    assert b"/FontFile2" in content
    assert b"Adobe-Japan1" not in content
    assert b"/Subtype /Image" in content
    assert len(re.findall(rb"/Type\s*/Page\b", content)) == 7


def test_builds_all_ci_artifacts_with_sha256_manifest(
    definition_file: Path, tmp_path: Path
) -> None:
    output_dir = tmp_path / "artifacts"

    manifest_path = build_artifacts(definition_file, output_dir)

    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    assert manifest["source"]["file"] == definition_file.name
    assert {item["format"] for item in manifest["artifacts"]} == {
        "html",
        "xlsx",
        "pdf",
        "mermaid",
        "svg",
        "png",
    }
    for item in manifest["artifacts"]:
        artifact = output_dir / item["file"]
        assert artifact.is_file()
        assert item["sha256"] == hashlib.sha256(artifact.read_bytes()).hexdigest()
